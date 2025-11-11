import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Cargar datos
df = pd.read_csv('../results/speedup_analysis.csv')

# Definir iteraciones de interés
iteraciones = [1000000, 10000000, 100000000]

# Obtener algoritmos y versiones
algoritmos = df['Algorithm'].unique()
versiones = df['Version'].unique()

with pd.ExcelWriter('../results/tablas_speedup_analysis.xlsx', engine='openpyxl') as writer:
    for algo in algoritmos:
        for version in versiones:
            datos = df[(df['Algorithm'] == algo) & (df['Version'] == version)]
            if datos.empty:
                continue
            # Pivot: filas=Threads, columnas=Iterations, valores=Speedup
            tabla = datos.pivot_table(index='Threads', columns='Iterations', values='Speedup', aggfunc='mean')
            cols_presentes = [col for col in iteraciones if col in tabla.columns]
            if not cols_presentes:
                continue
            tabla = tabla[cols_presentes]
            # Fila de promedios
            tabla.loc['Promedio'] = tabla.mean()
            # Escribir tabla en Excel
            sheet_name = f'{algo}_{version}'
            tabla.to_excel(writer, sheet_name=sheet_name, startrow=2)
            # Formato bonito con openpyxl
            wb = writer.book
            ws = wb[sheet_name]
            # Encabezado grande
            title = f'Speedup - {algo} {version}'
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1+len(cols_presentes))
            cell = ws.cell(row=1, column=1)
            cell.value = title
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Encabezados de columna en negrita y color
            for col in range(1, 2+len(cols_presentes)):
                ws.cell(row=3, column=col).font = Font(bold=True)
                ws.cell(row=3, column=col).fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
                ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
            # Bordes y ajuste de ancho
            thin = Side(border_style="thin", color="000000")
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1+len(cols_presentes)):
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for i in range(1, 2+len(cols_presentes)):
                ws.column_dimensions[get_column_letter(i)].width = 18

print("Tablas de speedup_analysis generadas en ../results/tablas_speedup_analysis.xlsx")
