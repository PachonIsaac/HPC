import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Cargar datos
benchmarks = pd.read_csv('../results/benchmarks.csv')

# Definir iteraciones de interés
iteraciones = [1000000, 10000000, 100000000]

# Obtener todas las versiones únicas
versiones = benchmarks['Version'].unique()
algoritmos = benchmarks['Algorithm'].unique()

with pd.ExcelWriter('../results/tablas_versiones.xlsx', engine='openpyxl') as writer:
    for version in versiones:
        for algo in algoritmos:
            datos = benchmarks[(benchmarks['Version'] == version) & (benchmarks['Algorithm'] == algo)]
            if datos.empty:
                continue
            # Pivot: filas=(Threads, Run), columnas=Iterations
            tabla = datos.pivot_table(index=['Threads', 'Run'], columns='Iterations', values='Time_Seconds', aggfunc='mean')
            cols_presentes = [col for col in iteraciones if col in tabla.columns]
            if not cols_presentes:
                continue
            tabla = tabla[cols_presentes]
            # Construir nueva tabla con promedios y speedup por Threads
            nueva_tabla = []
            serial = benchmarks[(benchmarks['Version'] == 'Serial') & (benchmarks['Algorithm'] == algo)]
            serial_prom = serial.groupby('Iterations')['Time_Seconds'].mean()
            for threads in sorted(tabla.index.get_level_values(0).unique()):
                sub = tabla.loc[threads]
                if isinstance(sub, pd.Series):
                    sub = sub.to_frame().T
                # Encabezado separador para los hilos
                encabezado = pd.DataFrame([[''] * len(cols_presentes)], columns=cols_presentes, index=[f'Threads = {threads}'])
                nueva_tabla.append(encabezado)
                # Renombrar filas solo como Prueba 1, 2, ...
                sub.index = [f'Prueba {i+1}' for i in range(len(sub))]
                nueva_tabla.append(sub)
                prom = sub.mean()
                prom_row = pd.DataFrame([prom], index=[f'Promedio'])
                nueva_tabla.append(prom_row)
                speedup = serial_prom[cols_presentes] / prom
                speedup_row = pd.DataFrame([speedup], index=[f'Speedup'])
                nueva_tabla.append(speedup_row)
            tabla_final = pd.concat(nueva_tabla)
            # Escribir tabla en Excel
            sheet_name = f'{algo}_{version}'
            tabla_final.to_excel(writer, sheet_name=sheet_name, startrow=2, header=True)
            # Formato bonito con openpyxl
            wb = writer.book
            ws = wb[sheet_name]
            title = f'Resultados - {algo} {version}'
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1+len(cols_presentes))
            cell = ws.cell(row=1, column=1)
            cell.value = title
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in range(1, 2+len(cols_presentes)):
                ws.cell(row=3, column=col).font = Font(bold=True)
                ws.cell(row=3, column=col).fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
                ws.cell(row=3, column=col).alignment = Alignment(horizontal='center')
            thin = Side(border_style="thin", color="000000")
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1+len(cols_presentes)):
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for i in range(1, 2+len(cols_presentes)):
                ws.column_dimensions[get_column_letter(i)].width = 18

print("Tablas generadas en ../results/tablas_versiones.xlsx")
